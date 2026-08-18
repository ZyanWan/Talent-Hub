#!/usr/bin/env python3
"""Build the resume evaluator workbook from JSON rows.

This script enforces the shared four-sheet core contract while allowing extra
fields. Headers are derived from dict rows or an explicit "_headers" entry.
"""

from __future__ import annotations

import argparse
import json
import math
import os
import sys
import tempfile
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

try:
    from workbook_contract import (
        CONCLUSION_FILL_COLORS,
        DEFAULT_HEADERS_PER_SHEET,
        EVIDENCE_SHEET,
        FREEZE_PANES,
        HEADER_FILL_COLOR,
        PHONE_SHEET,
        REQUIRED_SHEET_NAMES,
        STANDARD_MODULES,
        STANDARD_SHEET,
        SUMMARY_SHEET,
    )
except ModuleNotFoundError:  # Support python -m scripts.build_candidate_workbook
    from .workbook_contract import (
        CONCLUSION_FILL_COLORS,
        DEFAULT_HEADERS_PER_SHEET,
        EVIDENCE_SHEET,
        FREEZE_PANES,
        HEADER_FILL_COLOR,
        PHONE_SHEET,
        REQUIRED_SHEET_NAMES,
        STANDARD_MODULES,
        STANDARD_SHEET,
        SUMMARY_SHEET,
    )


# ---------------------------------------------------------------------------
# I/O helpers
# ---------------------------------------------------------------------------

def load_rows(path: Path) -> dict:
    data = json.loads(path.read_text(encoding="utf-8"))
    if not isinstance(data, dict):
        raise ValueError("Input JSON must be an object (sheet-name → rows).")
    return data


def resolve_headers(sheet_data: object, sheet_name: str) -> tuple[list[str], list[object]]:
    """Return (headers, rows) for a given sheet's data.

    The JSON may supply:
      - A list of dicts → headers come from dict keys (union of all keys).
      - A list of lists → look for "_headers" key in parent or fall back.
      - A dict with "_headers" + "_rows" → use those explicitly.
      - None / missing → empty.
    """
    # Explicit {"_headers": [...], "_rows": [...]} format
    if isinstance(sheet_data, dict):
        if "_headers" in sheet_data and "_rows" in sheet_data:
            return [str(value) for value in sheet_data["_headers"]], sheet_data["_rows"]
        # A regular dict is accepted as a single-row shorthand.
        return [str(k) for k in sheet_data.keys()], [sheet_data]

    if sheet_data is None:
        return [], []

    if not isinstance(sheet_data, list):
        raise ValueError(f"Sheet '{sheet_name}' data must be a list or dict.")

    if len(sheet_data) == 0:
        return DEFAULT_HEADERS_PER_SHEET.get(sheet_name, []), []

    first = sheet_data[0]
    if isinstance(first, dict):
        # Collect all keys across all rows to preserve order & completeness
        seen: list[str] = []
        for item in sheet_data:
            if isinstance(item, dict):
                for k in item:
                    if k not in seen and not k.startswith("_"):
                        seen.append(k)
        return seen, sheet_data

    # List of lists (or mixed) — use fallback headers
    return DEFAULT_HEADERS_PER_SHEET.get(sheet_name, []), sheet_data


def coerce_rows(rows: object, headers: list[str]) -> list[list[object]]:
    output: list[list[object]] = []
    for item in rows:
        if isinstance(item, dict):
            output.append([normalize_cell_value(item.get(h)) for h in headers])
        elif isinstance(item, list):
            output.append([normalize_cell_value(v) for v in item[: len(headers)]])
        else:
            output.append([normalize_cell_value(item)])
    return output


def normalize_cell_value(value: object) -> object:
    if value is None:
        return None
    if isinstance(value, bool):
        return value
    if isinstance(value, int):
        return value
    if isinstance(value, float):
        if not math.isfinite(value):
            raise ValueError("Excel cells cannot contain NaN or Infinity")
        return value
    if isinstance(value, str):
        return value
    if isinstance(value, (list, tuple)):
        return "\n".join(collection_item_text(item) for item in value)
    if isinstance(value, dict):
        return json.dumps(value, ensure_ascii=False, sort_keys=True)
    return str(value)


def collection_item_text(value: object) -> str:
    if value is None:
        return ""
    if isinstance(value, (dict, list, tuple)):
        return json.dumps(value, ensure_ascii=False, sort_keys=True)
    return str(value)


# ---------------------------------------------------------------------------
# Sheet writer
# ---------------------------------------------------------------------------

def write_literal_cell(cell, value: object) -> None:
    if isinstance(value, str):
        cell.value = value
        cell.data_type = "s"
        return
    cell.value = value


def write_sheet(ws, headers: list[str], rows: list[list[object]]) -> None:
    for column, header in enumerate(headers, start=1):
        write_literal_cell(ws.cell(row=1, column=column), str(header))
    for row_number, row in enumerate(rows, start=2):
        for column, value in enumerate(row, start=1):
            write_literal_cell(ws.cell(row=row_number, column=column), value)

    header_fill = PatternFill("solid", fgColor=HEADER_FILL_COLOR)
    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    ws.freeze_panes = FREEZE_PANES
    ws.auto_filter.ref = ws.dimensions

    conclusion_column = None
    if "结论" in headers:
        conclusion_column = headers.index("结论") + 1

    for row in ws.iter_rows(min_row=2):
        if conclusion_column:
            value = str(row[conclusion_column - 1].value or "")
            color = CONCLUSION_FILL_COLORS.get(value)
            if color:
                row[conclusion_column - 1].fill = PatternFill("solid", fgColor=color)

        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)

    for column_cells in ws.columns:
        max_length = 0
        column_letter = get_column_letter(column_cells[0].column)
        for cell in column_cells:
            value = "" if cell.value is None else str(cell.value)
            max_length = max(max_length, min(len(value), 40))
        ws.column_dimensions[column_letter].width = max(12, min(max_length + 2, 45))


# ---------------------------------------------------------------------------
# Main builder
# ---------------------------------------------------------------------------

def build_workbook(data: dict, output_path: Path) -> None:
    if output_path.suffix.lower() != ".xlsx":
        raise ValueError("Output path must use the .xlsx extension")
    missing_sheets = [name for name in REQUIRED_SHEET_NAMES if name not in data]
    if missing_sheets:
        raise ValueError(f"Input JSON missing required sheets: {', '.join(missing_sheets)}")

    wb = Workbook()
    ordered_sheet_names = list(REQUIRED_SHEET_NAMES) + [
        name for name in data if name not in REQUIRED_SHEET_NAMES
    ]

    for index, sheet_name in enumerate(ordered_sheet_names):
        sheet_data = data[sheet_name]
        headers, rows = resolve_headers(sheet_data, sheet_name)

        ws = wb.active if index == 0 else wb.create_sheet(title=sheet_name[:31])
        ws.title = sheet_name[:31]

        if headers or rows:  # Only write non-empty sheets
            coerced = coerce_rows(rows, headers) if rows else []
            write_sheet(ws, headers, coerced)

    output_path.parent.mkdir(parents=True, exist_ok=True)
    descriptor, temp_name = tempfile.mkstemp(
        prefix=f".{output_path.stem}.",
        suffix=output_path.suffix or ".xlsx",
        dir=str(output_path.parent),
    )
    os.close(descriptor)
    temp_path = Path(temp_name)
    try:
        wb.save(temp_path)
        os.replace(temp_path, output_path)
    finally:
        if temp_path.exists():
            temp_path.unlink()


# ---------------------------------------------------------------------------
# Sample payload (for --write-sample-json)
# ---------------------------------------------------------------------------

def sample_payload() -> dict:
    return {
        SUMMARY_SHEET: {
            "_headers": list(DEFAULT_HEADERS_PER_SHEET[SUMMARY_SHEET]),
            "_rows": [
                {
                    "推荐顺序": 1,
                    "候选人": "张三",
                    "结论": "B电话确认",
                    "一句话判定": "相关业务场景有线索，但核心动作和深度需电话确认。",
                    "当前/最近公司": "未体现",
                    "当前/最近岗位": "未体现",
                    "核心优势摘要": "未体现",
                    "关键风险/Blocker": "关键事实不清",
                    "下一步动作": "电话确认后再定",
                    "备注": "",
                }
            ],
        },
        EVIDENCE_SHEET: {
            "_headers": list(DEFAULT_HEADERS_PER_SHEET[EVIDENCE_SHEET]),
            "_rows": [
                {
                    "候选人": "张三",
                    "对象证据": "未体现",
                    "场景垂直性": "未体现",
                    "核心动作证据": "未体现",
                    "负责深度": "未体现",
                    "闭环证据": "未体现",
                    "工具/系统/证书": "未体现",
                    "规模/结果": "未体现",
                    "稳定性": "未体现",
                    "证据充分度": "低",
                }
            ],
        },
        PHONE_SHEET: [
            {
                "候选人": "张三",
                "优先级": "高",
                "确认焦点": "核心动作",
                "问题": "请说明一个由你独立推进并完成闭环的相关项目。",
                "当前证据": "简历未体现负责深度",
                "确认后影响": "B→A或B→C",
            }
        ],
        STANDARD_SHEET: [
            {"模块": module, "内容": f"{module}示例内容", "备注": ""}
            for module in STANDARD_MODULES
        ],
    }


def main() -> int:
    parser = argparse.ArgumentParser(description="Build resume evaluator workbook.")
    parser.add_argument("--input-json", type=Path, help="JSON rows keyed by sheet name")
    parser.add_argument("--output", type=Path, help="Output .xlsx path")
    parser.add_argument("--write-sample-json", type=Path, help="Write a sample JSON payload")
    args = parser.parse_args()

    if args.write_sample_json:
        args.write_sample_json.parent.mkdir(parents=True, exist_ok=True)
        args.write_sample_json.write_text(
            json.dumps(sample_payload(), ensure_ascii=False, indent=2),
            encoding="utf-8",
        )
        print(f"Sample written to {args.write_sample_json}")
        return 0

    if not args.input_json:
        parser.error("--input-json is required unless --write-sample-json is used")
    if not args.output:
        parser.error("--output is required when --input-json is used")

    build_workbook(load_rows(args.input_json), args.output)
    print(str(args.output))
    return 0


if __name__ == "__main__":
    sys.exit(main())
