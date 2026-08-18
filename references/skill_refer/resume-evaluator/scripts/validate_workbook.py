#!/usr/bin/env python3
"""Validate structural and security rules for resume evaluation workbooks."""

from __future__ import annotations

import argparse
import math
import sys
import zipfile
from pathlib import Path

from openpyxl import load_workbook

try:
    from workbook_contract import (
        CONCLUSION_FILL_COLORS,
        CONCLUSION_VALUES,
        CORE_FIELDS_PER_SHEET,
        EVIDENCE_SHEET,
        FREEZE_PANES,
        HEADER_FILL_COLOR,
        MIN_DATA_ROWS_PER_SHEET,
        OPTIONAL_STATUS_HEADERS,
        PHONE_SHEET,
        REQUIRED_SHEET_NAMES,
        ROW_KEY_FIELD_PER_SHEET,
        STANDARD_MODULES,
        STANDARD_SHEET,
        SUMMARY_SHEET,
        VALID_EVIDENCE_LEVELS,
        VALID_PHONE_PRIORITIES,
        VALID_STATUS_VALUES,
    )
except ModuleNotFoundError:  # Support python -m scripts.validate_workbook
    from .workbook_contract import (
        CONCLUSION_FILL_COLORS,
        CONCLUSION_VALUES,
        CORE_FIELDS_PER_SHEET,
        EVIDENCE_SHEET,
        FREEZE_PANES,
        HEADER_FILL_COLOR,
        MIN_DATA_ROWS_PER_SHEET,
        OPTIONAL_STATUS_HEADERS,
        PHONE_SHEET,
        REQUIRED_SHEET_NAMES,
        ROW_KEY_FIELD_PER_SHEET,
        STANDARD_MODULES,
        STANDARD_SHEET,
        SUMMARY_SHEET,
        VALID_EVIDENCE_LEVELS,
        VALID_PHONE_PRIORITIES,
        VALID_STATUS_VALUES,
    )


class WorkbookInspectionError(RuntimeError):
    """Raised when the workbook cannot be safely inspected."""


def normalized_text(value: object) -> str:
    return str(value or "").strip()


def header_map(ws) -> tuple[dict[str, int], list[str]]:
    headers: dict[str, int] = {}
    errors: list[str] = []
    for column in range(1, ws.max_column + 1):
        header = normalized_text(ws.cell(row=1, column=column).value)
        column_has_data = any(
            ws.cell(row=row, column=column).value not in (None, "")
            for row in range(2, ws.max_row + 1)
        )
        if not header:
            if column_has_data:
                errors.append(f"{ws.title} column {column} has data but no header")
            continue
        if header in headers:
            errors.append(f"{ws.title} has duplicate header: {header}")
            continue
        headers[header] = column
    return headers, errors


def row_has_any_data(ws, row_number: int) -> bool:
    return any(
        ws.cell(row=row_number, column=column).value not in (None, "")
        for column in range(1, ws.max_column + 1)
    )


def cell_value(ws, row_number: int, headers: dict[str, int], header: str) -> object:
    column = headers.get(header)
    if column is None:
        return None
    return ws.cell(row=row_number, column=column).value


def cell_text(ws, row_number: int, headers: dict[str, int], header: str) -> str:
    return normalized_text(cell_value(ws, row_number, headers, header))


def data_rows(ws, headers: dict[str, int], key_field: str) -> tuple[list[int], list[str]]:
    rows: list[int] = []
    errors: list[str] = []
    for row_number in range(2, ws.max_row + 1):
        if not row_has_any_data(ws, row_number):
            continue
        if not cell_text(ws, row_number, headers, key_field):
            errors.append(
                f"{ws.title} row {row_number} has data but missing row key: {key_field}"
            )
            continue
        rows.append(row_number)
    return rows, errors


def fill_rgb(cell) -> str:
    color = cell.fill.fgColor
    if color.type != "rgb" or not color.rgb:
        return ""
    return str(color.rgb)[-6:].upper()


def inspect_archive(path: Path) -> list[str]:
    errors: list[str] = []
    try:
        with zipfile.ZipFile(path) as archive:
            archive_names = archive.namelist()
            names = [name.lower() for name in archive_names]
            if any(name.endswith("vbaproject.bin") for name in names):
                errors.append("Workbook contains VBA or macros")
            if any(name.startswith("xl/externallinks/") for name in names):
                errors.append("Workbook contains external links")
            if "xl/connections.xml" in names:
                errors.append("Workbook contains external data connections")
            for archive_name in archive_names:
                if not archive_name.lower().endswith(".rels"):
                    continue
                relationship_xml = archive.read(archive_name).decode(
                    "utf-8", errors="ignore"
                )
                normalized_xml = relationship_xml.lower().replace("'", '"')
                if 'targetmode="external"' in normalized_xml:
                    errors.append("Workbook contains an external relationship")
                    break
    except (OSError, zipfile.BadZipFile) as exc:
        raise WorkbookInspectionError(f"Cannot inspect workbook archive: {exc}") from exc
    return errors


def load_for_validation(path: Path):
    if not path.exists():
        raise WorkbookInspectionError(f"Workbook does not exist: {path}")
    if not path.is_file():
        raise WorkbookInspectionError(f"Workbook path is not a file: {path}")
    if path.suffix.lower() != ".xlsx":
        raise WorkbookInspectionError("Workbook must use the .xlsx format")
    archive_errors = inspect_archive(path)
    try:
        workbook = load_workbook(path, data_only=False, keep_links=True)
    except Exception as exc:
        raise WorkbookInspectionError(f"Cannot open workbook: {exc}") from exc
    return workbook, archive_errors


def validate_loaded_workbook(wb, initial_errors: list[str]) -> tuple[list[str], list[str]]:
    errors = list(initial_errors)
    warnings: list[str] = []

    missing_sheets = [name for name in REQUIRED_SHEET_NAMES if name not in wb.sheetnames]
    errors.extend(f"Missing sheet: {name}" for name in missing_sheets)

    extra_sheets = [name for name in wb.sheetnames if name not in REQUIRED_SHEET_NAMES]
    warnings.extend(f"Unexpected extra sheet: {name}" for name in extra_sheets)

    if wb.sheetnames[: len(REQUIRED_SHEET_NAMES)] != list(REQUIRED_SHEET_NAMES):
        warnings.append(
            "Required sheets should appear in this order: " + ", ".join(REQUIRED_SHEET_NAMES)
        )
    if SUMMARY_SHEET in wb.sheetnames and wb.active.title != SUMMARY_SHEET:
        warnings.append(f"Active sheet should be {SUMMARY_SHEET}")

    for ws in wb.worksheets:
        for row in ws.iter_rows():
            for cell in row:
                if cell.data_type == "f":
                    errors.append(f"{ws.title}!{cell.coordinate} contains a formula")
                if cell.hyperlink is not None:
                    errors.append(f"{ws.title}!{cell.coordinate} contains a hyperlink")

    sheet_headers: dict[str, dict[str, int]] = {}
    sheet_rows: dict[str, list[int]] = {}

    for sheet_name in REQUIRED_SHEET_NAMES:
        if sheet_name not in wb.sheetnames:
            continue
        ws = wb[sheet_name]
        if ws.sheet_state != "visible":
            errors.append(f"{sheet_name} must be visible")

        headers, header_errors = header_map(ws)
        sheet_headers[sheet_name] = headers
        errors.extend(header_errors)

        missing_core = [
            field for field in CORE_FIELDS_PER_SHEET[sheet_name] if field not in headers
        ]
        if missing_core:
            errors.append(
                f"{sheet_name} missing core headers: {', '.join(missing_core)}"
            )

        if ws.freeze_panes != FREEZE_PANES:
            errors.append(f"{sheet_name} should freeze the first row at {FREEZE_PANES}")
        if not ws.auto_filter.ref:
            errors.append(f"{sheet_name} should enable autofilter")
        for merged_range in ws.merged_cells.ranges:
            errors.append(f"{sheet_name} has merged cells: {merged_range}")

        for cell in ws[1]:
            if cell.value in (None, ""):
                continue
            if not cell.font.bold or fill_rgb(cell) != HEADER_FILL_COLOR:
                warnings.append(f"{sheet_name} header style differs at {cell.coordinate}")

        key_field = ROW_KEY_FIELD_PER_SHEET[sheet_name]
        if key_field in headers:
            rows, row_errors = data_rows(ws, headers, key_field)
            errors.extend(row_errors)
        else:
            rows = []
        sheet_rows[sheet_name] = rows

        minimum_rows = MIN_DATA_ROWS_PER_SHEET[sheet_name]
        if len(rows) < minimum_rows:
            errors.append(
                f"{sheet_name} should contain at least {minimum_rows} data row(s)"
            )

        for row_number in rows:
            for field in CORE_FIELDS_PER_SHEET[sheet_name]:
                if field in headers and not cell_text(ws, row_number, headers, field):
                    errors.append(
                        f"{sheet_name} row {row_number} missing core field: {field}"
                    )

    summary_candidates: set[str] = set()
    summary_conclusions: dict[str, str] = {}
    b_candidates: set[str] = set()

    if SUMMARY_SHEET in sheet_headers:
        ws = wb[SUMMARY_SHEET]
        headers = sheet_headers[SUMMARY_SHEET]
        ranks: set[int] = set()
        for row_number in sheet_rows.get(SUMMARY_SHEET, []):
            candidate = cell_text(ws, row_number, headers, "候选人")
            conclusion = cell_text(ws, row_number, headers, "结论")
            if candidate in summary_candidates:
                warnings.append(
                    f"{SUMMARY_SHEET} has duplicate candidate name: {candidate}"
                )
            summary_candidates.add(candidate)
            summary_conclusions[candidate] = conclusion

            if conclusion not in CONCLUSION_VALUES:
                errors.append(
                    f"{SUMMARY_SHEET} row {row_number} invalid 结论: {conclusion}"
                )
            elif conclusion == "B电话确认":
                b_candidates.add(candidate)

            rank = cell_value(ws, row_number, headers, "推荐顺序")
            rank_number: int | None = None
            if isinstance(rank, int) and not isinstance(rank, bool) and rank > 0:
                rank_number = rank
            elif (
                isinstance(rank, float)
                and math.isfinite(rank)
                and rank.is_integer()
                and rank > 0
            ):
                rank_number = int(rank)

            if rank_number is None:
                errors.append(
                    f"{SUMMARY_SHEET} row {row_number} 推荐顺序 must be a positive integer"
                )
            else:
                if rank_number in ranks:
                    errors.append(f"{SUMMARY_SHEET} duplicate 推荐顺序: {rank_number}")
                ranks.add(rank_number)

            if conclusion in CONCLUSION_FILL_COLORS and "结论" in headers:
                cell = ws.cell(row=row_number, column=headers["结论"])
                expected_color = CONCLUSION_FILL_COLORS[conclusion]
                if fill_rgb(cell) != expected_color:
                    warnings.append(
                        f"{SUMMARY_SHEET}!{cell.coordinate} conclusion color differs"
                    )

    evidence_candidates: set[str] = set()
    if EVIDENCE_SHEET in sheet_headers:
        ws = wb[EVIDENCE_SHEET]
        headers = sheet_headers[EVIDENCE_SHEET]
        for row_number in sheet_rows.get(EVIDENCE_SHEET, []):
            candidate = cell_text(ws, row_number, headers, "候选人")
            evidence_candidates.add(candidate)
            if candidate not in summary_candidates:
                errors.append(f"{EVIDENCE_SHEET} has unknown candidate: {candidate}")

            level = cell_text(ws, row_number, headers, "证据充分度")
            if level and level not in VALID_EVIDENCE_LEVELS:
                errors.append(
                    f"{EVIDENCE_SHEET} row {row_number} invalid 证据充分度: {level}"
                )
            for field in OPTIONAL_STATUS_HEADERS:
                if field not in headers:
                    continue
                status = cell_text(ws, row_number, headers, field)
                if status and status not in VALID_STATUS_VALUES:
                    errors.append(
                        f"{EVIDENCE_SHEET} row {row_number} invalid {field}: {status}"
                    )

    for candidate in sorted(summary_candidates - evidence_candidates):
        errors.append(f"{EVIDENCE_SHEET} missing candidate: {candidate}")

    phone_candidates: set[str] = set()
    if PHONE_SHEET in sheet_headers:
        ws = wb[PHONE_SHEET]
        headers = sheet_headers[PHONE_SHEET]
        for row_number in sheet_rows.get(PHONE_SHEET, []):
            candidate = cell_text(ws, row_number, headers, "候选人")
            phone_candidates.add(candidate)
            if candidate not in summary_candidates:
                errors.append(f"{PHONE_SHEET} has unknown candidate: {candidate}")

            priority = cell_text(ws, row_number, headers, "优先级")
            if priority and priority not in VALID_PHONE_PRIORITIES:
                errors.append(
                    f"{PHONE_SHEET} row {row_number} invalid 优先级: {priority}"
                )

    for candidate in sorted(b_candidates - phone_candidates):
        errors.append(f"{PHONE_SHEET} missing question for B candidate: {candidate}")
    for candidate in sorted(phone_candidates - b_candidates):
        conclusion = summary_conclusions.get(candidate, "")
        if conclusion:
            warnings.append(
                f"{PHONE_SHEET} contains {conclusion} candidate: {candidate}"
            )

    if STANDARD_SHEET in sheet_headers:
        ws = wb[STANDARD_SHEET]
        headers = sheet_headers[STANDARD_SHEET]
        modules = {
            cell_text(ws, row_number, headers, "模块")
            for row_number in sheet_rows.get(STANDARD_SHEET, [])
        }
        for module in STANDARD_MODULES:
            if module not in modules:
                errors.append(f"{STANDARD_SHEET} missing module: {module}")

    return errors, warnings


def validate_workbook_detailed(path: Path) -> tuple[list[str], list[str]]:
    wb, archive_errors = load_for_validation(path)
    try:
        return validate_loaded_workbook(wb, archive_errors)
    finally:
        wb.close()


def validate_workbook(path: Path) -> list[str]:
    """Compatibility wrapper returning only blocking validation errors."""
    try:
        errors, _ = validate_workbook_detailed(path)
        return errors
    except WorkbookInspectionError as exc:
        return [str(exc)]


def main() -> int:
    parser = argparse.ArgumentParser(description="Validate resume evaluator workbook.")
    parser.add_argument("workbook", type=Path)
    parser.add_argument(
        "--warnings-as-errors",
        action="store_true",
        help="Fail when non-blocking presentation warnings are present",
    )
    args = parser.parse_args()

    try:
        errors, warnings = validate_workbook_detailed(args.workbook)
    except WorkbookInspectionError as exc:
        print(f"ERROR: {exc}", file=sys.stderr)
        return 2

    for warning in warnings:
        print(f"WARNING: {warning}", file=sys.stderr)
    for error in errors:
        print(f"ERROR: {error}", file=sys.stderr)

    if errors or (args.warnings_as_errors and warnings):
        return 1

    print("PASS: workbook satisfies resume-evaluator structural and security contract")
    print("This does not validate the correctness of screening criteria or A/B/C decisions.")
    return 0


if __name__ == "__main__":
    sys.exit(main())
