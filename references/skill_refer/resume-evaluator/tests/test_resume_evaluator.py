from __future__ import annotations

import importlib
import sys
import tempfile
import unittest
import zipfile
from pathlib import Path

from openpyxl import load_workbook


SKILL_DIR = Path(__file__).resolve().parents[1]
SCRIPTS_DIR = SKILL_DIR / "scripts"
sys.path.insert(0, str(SCRIPTS_DIR))

builder = importlib.import_module("build_candidate_workbook")
extractor = importlib.import_module("extract_resume_text")
validator = importlib.import_module("validate_workbook")
contract = importlib.import_module("workbook_contract")

EVIDENCE_SHEET = contract.EVIDENCE_SHEET
PHONE_SHEET = contract.PHONE_SHEET
STANDARD_SHEET = contract.STANDARD_SHEET
SUMMARY_SHEET = contract.SUMMARY_SHEET


class ResumeEvaluatorTests(unittest.TestCase):
    def test_same_stem_outputs_are_unique_and_stable(self) -> None:
        with tempfile.TemporaryDirectory() as temp_dir:
            root = Path(temp_dir)
            source_a = root / "a" / "same.txt"
            source_b = root / "b" / "same.md"
            source_a.parent.mkdir()
            source_b.parent.mkdir()
            source_a.write_text("first", encoding="utf-8")
            source_b.write_text("second", encoding="utf-8")
            output_dir = root / "parsed_text"

            records = [
                {"file": str(source_a), "text": "first"},
                {"file": str(source_b), "text": "second"},
            ]
            extractor.write_outputs(records, output_dir)
            first_paths = [Path(record["text_path"]) for record in records]

            self.assertNotEqual(first_paths[0], first_paths[1])
            self.assertEqual(first_paths[0].read_text(encoding="utf-8"), "first")
            self.assertEqual(first_paths[1].read_text(encoding="utf-8"), "second")

            repeated = [
                {"file": str(source_a), "text": "first"},
                {"file": str(source_b), "text": "second"},
            ]
            extractor.write_outputs(repeated, output_dir)
            self.assertEqual(
                first_paths,
                [Path(record["text_path"]) for record in repeated],
            )

    def test_builder_writes_formula_like_text_as_strings_and_preserves_numbers(self) -> None:
        attacks = {
            "攻击值1": "=HYPERLINK(\"https://example.invalid\",\"x\")",
            "攻击值2": "+CMD",
            "攻击值3": "-1+1",
            "攻击值4": "@SUM(A1:A2)",
            "攻击值5": " \t=1+1",
        }
        malicious_header = "=HYPERLINK(\"https://example.invalid\",\"header\")"
        payload = builder.sample_payload()
        summary_row = payload[SUMMARY_SHEET]["_rows"][0]
        summary_row.update(attacks)
        summary_row[malicious_header] = "ordinary text"
        summary_row["推荐顺序"] = 10
        payload[SUMMARY_SHEET]["_headers"].extend([*attacks, malicious_header])

        with tempfile.TemporaryDirectory() as temp_dir:
            output = Path(temp_dir) / "output.xlsx"
            builder.build_workbook(payload, output)
            wb = load_workbook(output, data_only=False)
            ws = wb[SUMMARY_SHEET]
            headers = [cell.value for cell in ws[1]]

            rank = ws.cell(2, headers.index("推荐顺序") + 1)
            self.assertEqual(rank.value, 10)
            self.assertEqual(rank.data_type, "n")

            for header, expected in attacks.items():
                cell = ws.cell(2, headers.index(header) + 1)
                self.assertEqual(cell.value, expected)
                self.assertEqual(cell.data_type, "s")

            header_cell = ws.cell(1, headers.index(malicious_header) + 1)
            self.assertEqual(header_cell.value, malicious_header)
            self.assertEqual(header_cell.data_type, "s")

            errors, warnings = validator.validate_workbook_detailed(output)
            self.assertEqual(errors, [])
            self.assertEqual(warnings, [])

    def test_sample_workbook_satisfies_contract(self) -> None:
        with tempfile.TemporaryDirectory() as temp_dir:
            output = Path(temp_dir) / "sample.xlsx"
            builder.build_workbook(builder.sample_payload(), output)
            wb = load_workbook(output, data_only=False)
            self.assertEqual(
                wb.sheetnames[:4],
                [SUMMARY_SHEET, EVIDENCE_SHEET, PHONE_SHEET, STANDARD_SHEET],
            )
            errors, warnings = validator.validate_workbook_detailed(output)
            self.assertEqual(errors, [])
            self.assertEqual(warnings, [])

    def test_validator_rejects_formula_and_invalid_conclusion(self) -> None:
        payload = builder.sample_payload()
        payload[SUMMARY_SHEET]["_rows"][0]["结论"] = "Apple"

        with tempfile.TemporaryDirectory() as temp_dir:
            output = Path(temp_dir) / "invalid.xlsx"
            builder.build_workbook(payload, output)
            wb = load_workbook(output)
            ws = wb[SUMMARY_SHEET]
            headers = [cell.value for cell in ws[1]]
            ws.cell(2, headers.index("一句话判定") + 1).value = "=1+1"
            wb.save(output)

            errors, _ = validator.validate_workbook_detailed(output)
            self.assertTrue(any("invalid 结论: Apple" in error for error in errors))
            self.assertTrue(any("contains a formula" in error for error in errors))

    def test_validator_rejects_formula_in_extra_sheet(self) -> None:
        with tempfile.TemporaryDirectory() as temp_dir:
            output = Path(temp_dir) / "extra-formula.xlsx"
            builder.build_workbook(builder.sample_payload(), output)
            wb = load_workbook(output)
            extra = wb.create_sheet("附加信息")
            extra["A1"] = "=1+1"
            wb.save(output)

            errors, _ = validator.validate_workbook_detailed(output)
            self.assertTrue(
                any("附加信息!A1 contains a formula" in error for error in errors)
            )

    def test_validator_rejects_external_connection_part(self) -> None:
        with tempfile.TemporaryDirectory() as temp_dir:
            output = Path(temp_dir) / "connection.xlsx"
            builder.build_workbook(builder.sample_payload(), output)
            with zipfile.ZipFile(output, "a") as archive:
                archive.writestr("xl/connections.xml", "<connections />")

            errors, _ = validator.validate_workbook_detailed(output)
            self.assertTrue(
                any("external data connections" in error for error in errors)
            )

    def test_validator_reports_missing_header_without_crashing(self) -> None:
        with tempfile.TemporaryDirectory() as temp_dir:
            output = Path(temp_dir) / "missing-header.xlsx"
            builder.build_workbook(builder.sample_payload(), output)
            wb = load_workbook(output)
            ws = wb[SUMMARY_SHEET]
            headers = [cell.value for cell in ws[1]]
            ws.delete_cols(headers.index("结论") + 1)
            wb.save(output)

            errors, _ = validator.validate_workbook_detailed(output)
            self.assertTrue(any("missing core headers: 结论" in error for error in errors))

    def test_validator_requires_phone_question_for_b_candidate(self) -> None:
        payload = builder.sample_payload()
        payload[PHONE_SHEET] = []

        with tempfile.TemporaryDirectory() as temp_dir:
            output = Path(temp_dir) / "missing-phone.xlsx"
            builder.build_workbook(payload, output)
            errors, _ = validator.validate_workbook_detailed(output)
            self.assertTrue(
                any("missing question for B candidate" in error for error in errors)
            )

    def test_validator_rejects_string_recommendation_order(self) -> None:
        payload = builder.sample_payload()
        payload[SUMMARY_SHEET]["_rows"][0]["推荐顺序"] = "10"

        with tempfile.TemporaryDirectory() as temp_dir:
            output = Path(temp_dir) / "string-rank.xlsx"
            builder.build_workbook(payload, output)
            errors, _ = validator.validate_workbook_detailed(output)
            self.assertTrue(
                any("推荐顺序 must be a positive integer" in error for error in errors)
            )

    def test_validator_closes_workbook_before_same_path_rebuild(self) -> None:
        with tempfile.TemporaryDirectory() as temp_dir:
            output = Path(temp_dir) / "rebuild.xlsx"
            payload = builder.sample_payload()
            builder.build_workbook(payload, output)
            errors, warnings = validator.validate_workbook_detailed(output)
            self.assertEqual(errors, [])
            self.assertEqual(warnings, [])
            builder.build_workbook(payload, output)
            self.assertTrue(output.exists())

    def test_builder_requires_all_four_sheets(self) -> None:
        payload = builder.sample_payload()
        del payload[STANDARD_SHEET]
        with tempfile.TemporaryDirectory() as temp_dir:
            with self.assertRaisesRegex(ValueError, "missing required sheets"):
                builder.build_workbook(payload, Path(temp_dir) / "missing.xlsx")


if __name__ == "__main__":
    unittest.main()
