from __future__ import annotations

from datetime import date, datetime, time, timedelta
from pathlib import Path
from zipfile import BadZipFile

from openpyxl import load_workbook
from openpyxl.utils.exceptions import InvalidFileException


MAX_MARKDOWN_CHARS = 1_000_000
MAX_WORKBOOK_ROWS = 500
MAX_WORKBOOK_COLUMNS = 40
MAX_CELL_CHARS = 4_000


def resolve_artifact(job_dir: Path, filename: str) -> Path:
    if not filename:
        raise FileNotFoundError(filename)
    root = job_dir.resolve()
    path = (root / filename).resolve()
    if path.parent != root:
        raise ValueError("任务文件路径无效。")
    if not path.is_file():
        raise FileNotFoundError(filename)
    return path


def markdown_preview(path: Path) -> dict[str, object]:
    with path.open("r", encoding="utf-8", errors="replace") as handle:
        content = handle.read(MAX_MARKDOWN_CHARS + 1)
    truncated = len(content) > MAX_MARKDOWN_CHARS
    return {
        "kind": "markdown",
        "filename": path.name,
        "content": content[:MAX_MARKDOWN_CHARS],
        "truncated": truncated,
    }


def _preview_value(value: object) -> tuple[object, bool]:
    if value is None:
        return "", False
    if isinstance(value, (datetime, date, time)):
        return value.isoformat(), False
    if isinstance(value, timedelta):
        return str(value), False
    if isinstance(value, (bool, int, float)):
        return value, False
    text = str(value)
    if len(text) > MAX_CELL_CHARS:
        return f"{text[:MAX_CELL_CHARS]}…", True
    return text, False


def workbook_preview(path: Path) -> dict[str, object]:
    try:
        workbook = load_workbook(path, read_only=True, data_only=True, keep_links=False)
    except (OSError, BadZipFile, InvalidFileException, KeyError) as exc:
        raise ValueError("无法读取评估表格预览。") from exc

    sheets: list[dict[str, object]] = []
    try:
        for worksheet in workbook.worksheets:
            total_rows = int(worksheet.max_row or 0)
            total_columns = int(worksheet.max_column or 0)
            shown_rows = min(total_rows, MAX_WORKBOOK_ROWS)
            shown_columns = min(total_columns, MAX_WORKBOOK_COLUMNS)
            rows: list[list[object]] = []
            truncated = total_rows > shown_rows or total_columns > shown_columns
            if shown_rows and shown_columns:
                for source_row in worksheet.iter_rows(
                    min_row=1,
                    max_row=shown_rows,
                    min_col=1,
                    max_col=shown_columns,
                    values_only=True,
                ):
                    row: list[object] = []
                    for source_value in source_row:
                        value, value_truncated = _preview_value(source_value)
                        truncated = truncated or value_truncated
                        row.append(value)
                    rows.append(row)
            sheets.append({
                "name": worksheet.title,
                "rows": rows,
                "total_rows": total_rows,
                "total_columns": total_columns,
                "truncated": truncated,
            })
    finally:
        workbook.close()

    return {
        "kind": "workbook",
        "filename": path.name,
        "sheets": sheets,
        "truncated": any(bool(sheet["truncated"]) for sheet in sheets),
    }
